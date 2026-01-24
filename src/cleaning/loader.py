"""File loading utilities with encoding detection and format handling."""

from pathlib import Path
from typing import Dict, Any, Tuple, Optional
import pandas as pd
import chardet

from .utils import get_file_extension


class LoadError(Exception):
    """Raised when a file cannot be loaded."""
    pass


def detect_encoding(filepath: str) -> str:
    """
    Detect the encoding of a file using chardet.

    Args:
        filepath: Path to the file

    Returns:
        Detected encoding string (e.g., 'utf-8', 'latin-1')
    """
    # Read first 100KB for detection (sufficient for most files)
    with open(filepath, 'rb') as f:
        raw_data = f.read(100000)

    result = chardet.detect(raw_data)
    encoding = result.get('encoding', 'utf-8')

    # Handle common chardet quirks
    if encoding is None:
        encoding = 'utf-8'
    elif encoding.lower() == 'ascii':
        # ASCII is a subset of UTF-8, use UTF-8 for better compatibility
        encoding = 'utf-8'

    return encoding


def load_csv(
    filepath: str,
    encoding: Optional[str] = None,
    **kwargs
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Load a CSV file with automatic encoding detection.

    Args:
        filepath: Path to the CSV file
        encoding: Optional encoding override. If None, auto-detect.
        **kwargs: Additional arguments passed to pd.read_csv

    Returns:
        Tuple of (DataFrame, metadata dict)

    Raises:
        LoadError: If the file cannot be loaded
    """
    path = Path(filepath)

    if not path.exists():
        raise LoadError(f"File not found: {filepath}")

    # Detect encoding if not provided
    if encoding is None:
        encoding = detect_encoding(filepath)

    metadata = {
        'source_file': str(path.absolute()),
        'file_type': 'csv',
        'detected_encoding': encoding,
        'encoding_override': encoding if encoding else None,
    }

    try:
        # Try with detected encoding
        df = pd.read_csv(filepath, encoding=encoding, **kwargs)
    except UnicodeDecodeError:
        # Fallback encodings to try
        fallback_encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        df = None

        for fallback in fallback_encodings:
            if fallback == encoding:
                continue
            try:
                df = pd.read_csv(filepath, encoding=fallback, **kwargs)
                metadata['detected_encoding'] = fallback
                metadata['encoding_fallback_used'] = True
                break
            except UnicodeDecodeError:
                continue

        if df is None:
            raise LoadError(
                f"Could not decode file with any encoding. "
                f"Tried: {encoding}, {', '.join(fallback_encodings)}"
            )

    metadata['original_rows'] = len(df)
    metadata['original_columns'] = len(df.columns)
    metadata['column_names'] = list(df.columns)

    return df, metadata


def load_excel(
    filepath: str,
    sheet_name: Optional[str] = None,
    handle_merged_cells: bool = True,
    **kwargs
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Load an Excel file with merged cell handling.

    Args:
        filepath: Path to the Excel file
        sheet_name: Optional sheet name. If None, loads first sheet.
        handle_merged_cells: Whether to forward-fill merged cells
        **kwargs: Additional arguments passed to pd.read_excel

    Returns:
        Tuple of (DataFrame, metadata dict)

    Raises:
        LoadError: If the file cannot be loaded
    """
    path = Path(filepath)

    if not path.exists():
        raise LoadError(f"File not found: {filepath}")

    metadata = {
        'source_file': str(path.absolute()),
        'file_type': 'excel',
        'sheet_name': sheet_name,
    }

    try:
        # Load the workbook to inspect merged cells
        from openpyxl import load_workbook

        wb = load_workbook(filepath, data_only=True)
        sheet_names = wb.sheetnames
        metadata['available_sheets'] = sheet_names

        # Determine which sheet to use
        if sheet_name is None:
            sheet_name = sheet_names[0]
            metadata['sheet_name'] = sheet_name
        elif sheet_name not in sheet_names:
            raise LoadError(
                f"Sheet '{sheet_name}' not found. "
                f"Available sheets: {', '.join(sheet_names)}"
            )

        ws = wb[sheet_name]

        # Detect merged cell ranges
        merged_ranges = list(ws.merged_cells.ranges)
        metadata['merged_cell_count'] = len(merged_ranges)
        metadata['merged_ranges'] = [str(r) for r in merged_ranges]

        wb.close()

    except ImportError:
        raise LoadError("openpyxl is required to load Excel files")
    except Exception as e:
        raise LoadError(f"Error inspecting Excel file: {e}")

    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name, **kwargs)
    except Exception as e:
        raise LoadError(f"Error reading Excel file: {e}")

    # Handle merged cells by forward-filling
    if handle_merged_cells and metadata['merged_cell_count'] > 0:
        # Forward fill for merged cells in both directions
        df = df.ffill(axis=0)  # Fill down (rows)
        # Note: We don't fill right because merged cells typically span rows
        metadata['merged_cells_handled'] = True

    metadata['original_rows'] = len(df)
    metadata['original_columns'] = len(df.columns)
    metadata['column_names'] = list(df.columns)

    return df, metadata


def load_file(
    filepath: str,
    **kwargs
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    Load a data file (CSV or Excel) with automatic format detection.

    This is the main entry point for loading files. It detects the file
    format based on extension and delegates to the appropriate loader.

    Args:
        filepath: Path to the data file
        **kwargs: Additional arguments passed to the specific loader

    Returns:
        Tuple of (DataFrame, metadata dict)

    Raises:
        LoadError: If the file cannot be loaded or format is unsupported

    Example:
        >>> df, meta = load_file("data.csv")
        >>> print(f"Loaded {meta['original_rows']} rows")
    """
    extension = get_file_extension(filepath)

    if extension == 'csv':
        return load_csv(filepath, **kwargs)
    elif extension in ('xlsx', 'xls'):
        return load_excel(filepath, **kwargs)
    else:
        raise LoadError(
            f"Unsupported file format: .{extension}. "
            f"Supported formats: .csv, .xlsx, .xls"
        )
